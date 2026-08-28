#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成主表模板：55 字段 CSV + XLSX（含示例行与数据校验下拉由后续阶段补充）"""
import csv, os

BASE = "/Users/fanglu/WorkBuddy/2026-08-28-13-11-43/全国肝病联盟与2030专题知识库/01_主表"

# 55 字段（按用户给定顺序，附组别与说明）
FIELDS = [
    ("文献唯一标识", "A"), ("PMID", "A"), ("DOI", "A"), ("中文标题", "B"), ("英文标题", "B"),
    ("发布年份", "B"), ("发布日期", "B"), ("期刊或发布机构", "B"), ("第一作者", "B"), ("通讯作者", "B"),
    ("作者单位", "B"), ("国家/地区", "B"), ("是否中国研究", "C"), ("是否中国患者", "C"), ("是否中国多中心研究", "C"),
    ("文献类型", "B"), ("证据等级", "D"), ("研究设计", "D"), ("研究阶段", "D"), ("样本量", "D"),
    ("研究人群", "D"), ("初治/经治", "E"), ("HBeAg状态", "E"), ("是否肝硬化", "E"), ("治疗方案", "F"),
    ("对照方案", "F"), ("治疗时间", "F"), ("随访时间", "F"), ("基线HBsAg", "G"), ("基线HBV DNA", "G"),
    ("HBsAg下降速度或变化幅度", "H"), ("HBV DNA下降速度或变化幅度", "H"), ("主要终点", "H"), ("HBsAg清除率", "H"),
    ("HBeAg血清学转换率", "H"), ("病毒学应答率", "H"), ("生化学应答率", "H"), ("肝癌发生或复发结局", "H"),
    ("安全性", "H"), ("核心结论", "I"), ("临床价值", "I"), ("对中国实践的启示", "I"), ("对2030行动的意义", "I"),
    ("对全国肝病联盟的意义", "I"), ("所属筛诊治管康环节", "I"), ("所属一级专题", "I"), ("所属二级专题", "I"),
    ("适用患者分层", "E"), ("当前争议", "J"), ("研究局限", "J"), ("是否建议重点阅读", "J"),
    ("建议关注优先级", "J"), ("原文链接", "A"), ("二次解读链接", "A"), ("最后核查日期", "A"),
]
assert len(FIELDS) == 55, len(FIELDS)

GROUP = {"A": "A_标识与溯源", "B": "B_题录信息", "C": "C_中国属性", "D": "D_方法与证据",
         "E": "E_人群特征", "F": "F_治疗方案", "G": "G_基线指标", "H": "H_疗效与结局",
         "I": "I_价值与定位", "J": "J_质量与分级"}

headers = [f"{f}({GROUP[g]})" if g in ("A",) else f for f, g in FIELDS]  # 组别写入表头便于识别
# 更清晰：表头带组别前缀
headers = [f"[{GROUP[g]}] {f}" for f, g in FIELDS]

example = [
    "KB-0001", "待核查", "待核查", "示例：核苷经治患者序贯聚乙二醇干扰素α的HBsAg清除率（样例卡，非真实录入）", "",
    "2025", "2025-06-10", "EASL 2025 会议摘要（示例）", "", "", "", "中国", "是", "是", "待核查",
    "会议摘要", "A", "多中心队列（示例）", "不适用", "待核查", "核苷经治CHB患者（示例）",
    "经治", "混合", "待核查", "NAs联合/序贯Peg-IFNα（示例）", "NAs单药", "96周", "96周",
    "待核查", "待核查", "待核查", "待核查", "48周HBsAg清除率（示例）", "待核查", "待核查", "待核查", "待核查",
    "待核查", "待核查", "待核查", "待核查", "待核查", "待核查", "筛-诊-治-管-康:治", "T4B", "4B-3",
    "核苷经治HBsAg低水平（示例）", "待核查", "待核查", "是", "P0", "", "", "2026-08-28",
]

csv_path = os.path.join(BASE, "主表模板_55字段.csv")
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(headers)
    w.writerow(example)

# XLSX（若 openpyxl 可用）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "主表"
    ws.append(headers)
    ws.append(example)
    fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[1]:
        c.font = Font(bold=True, color="1F4E79")
        c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for col in ws.columns:
        width = max(14, min(40, max((len(str(c.value or "")) for c in col)) + 2))
        ws.column_dimensions[col[0].column_letter].width = width
    ws.freeze_panes = "A2"
    xlsx_path = os.path.join(BASE, "主表模板_55字段.xlsx")
    wb.save(xlsx_path)
    print("OK: csv + xlsx 均已生成")
except ImportError:
    print("OK: csv 已生成（openpyxl 不可用，xlsx 跳过）")

print("表头数:", len(headers))
